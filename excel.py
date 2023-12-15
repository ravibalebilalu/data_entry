from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from tabulate import tabulate

class ExcelHandler:
    def __init__(self, file_name):
        self.file_name = file_name

    def write_table(self, data, sheet_name='Sheet', table_name='Table1'):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

        for row in data:
            sheet.append(row)

        table = Table(displayName=table_name, ref=sheet.dimensions)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style
        sheet.add_table(table)

        workbook.save(self.file_name)

    def read_table(self, sheet_name='Sheet'):
        workbook = load_workbook(filename=self.file_name)
        sheet = workbook[sheet_name]

        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        return data