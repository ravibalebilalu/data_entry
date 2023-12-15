from excel import ExcelHandler
from openpyxl import Workbook,load_workbook
from tabulate import tabulate
import pyinputplus as pyip
import os
from logger import logging
logging.info("Script started")
files = [f[:-5] for f in os.listdir() if os.path.isfile(f) and f.endswith(".xlsx")]
print("List of files : ",files)
file_name_query = pyip.inputStr("Enter the file name : ")
file_name = file_name_query + ".xlsx"
#check for file
if not os.path.exists(file_name):
    print("No file found! Let us create one")
    workbook = Workbook()
    column_names = []
    sheet = workbook.active
    #ask user to enter number of columns and column names
    col_count = pyip.inputInt("Enter the number of columns : ")
    for i in range(col_count):
        name = pyip.inputStr("Enter the column name " + str(i + 1) + " ")
        column_names.append(name)
    excel = ExcelHandler(file_name)
    excel.write_table([column_names])
    logging.info(file_name + " created")

    # if file exists then read file
excel = ExcelHandler(file_name)
workbook = load_workbook(file_name)
sheet = workbook.active
logging.info(file_name + " read")

#column names
header_row = sheet[1]
column_names = [cell.value for cell in header_row]
print(column_names)
#sheet names
available_sheets = workbook.sheetnames
selection = pyip.inputStr("Select [w] for entry or [r] for read : ")
if selection == "r":
    print(tabulate(excel.read_table(available_sheets[0])[1:],headers = column_names,tablefmt = "grid"))
elif selection == "w":
    print("Let us enter the data !")
    print("Colum names are : ", column_names)
    entry_flag = "!q"
    while entry_flag != "q":
        data_list = []
        for i in range(len(column_names)):
            entry = pyip.inputStr("Enter " + column_names[i] + " : ")
            data_list.append(entry)
        existing_data = excel.read_table(available_sheets[0])
        existing_data.extend([data_list])
        excel.write_table(existing_data)
        entry_flag = pyip.inputStr("Enter 'c' for continue or 'q' for exit : ")
    logging.info("Data apended")
task = pyip.inputStr("Would You like display table? [y] or [n] :")
if task == "y":
    print(tabulate(excel.read_table(available_sheets[0])[1:],headers = column_names,tablefmt = "grid"))
    print("Finishing the work\n HAVE A NICE DAY!")
elif task == "n":
    print("Finishing the work\n HAVE A NICE DAY!")
logging.info("Script ended")